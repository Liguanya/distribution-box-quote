"""
Excel报价单生成模块
根据元器件数据生成专业的Excel报价单
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


def generate_excel(data, boxes, output_dir=None):
    """
    生成Excel报价单
    
    Args:
        data: 所有元器件数据列表
        boxes: 按配电箱分组的数据
        output_dir: 输出目录
    
    Returns:
        str: 生成的Excel文件路径
    """
    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 生成文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"配电箱报价单_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # 创建工作簿
    wb = Workbook()
    
    # 样式定义
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E88E5', end_color='1E88E5', fill_type='solid')
    title_font = Font(name='微软雅黑', size=14, bold=True)
    normal_font = Font(name='微软雅黑', size=10)
    money_font = Font(name='微软雅黑', size=10, bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # ========== 汇总表 ==========
    ws_summary = wb.active
    ws_summary.title = "报价汇总"
    
    # 标题
    ws_summary['A1'] = "配电箱报价汇总表"
    ws_summary['A1'].font = Font(name='微软雅黑', size=16, bold=True, color='1E88E5')
    ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.merge_cells('A1:F1')
    
    # 报价日期
    ws_summary['A2'] = f"报价日期：{datetime.now().strftime('%Y年%m月%d日')}"
    ws_summary['A2'].font = normal_font
    ws_summary.merge_cells('A2:F2')
    
    # 表头
    headers = ['序号', '配电箱名称', '箱体尺寸', '元器件数量', '报价合计(元)', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # 数据行
    row = 5
    total_amount = 0
    total_boxes = 0
    
    for idx, (box_name, box_info) in enumerate(boxes.items(), 1):
        box_total = sum(c['total'] for c in box_info['components'])
        total_amount += box_total
        total_boxes += 1
        
        ws_summary.cell(row=row, column=1, value=idx).alignment = center_align
        ws_summary.cell(row=row, column=2, value=box_name).alignment = left_align
        ws_summary.cell(row=row, column=3, value=box_info['size']).alignment = center_align
        ws_summary.cell(row=row, column=4, value=len(box_info['components'])).alignment = center_align
        
        cell_total = ws_summary.cell(row=row, column=5, value=box_total)
        cell_total.number_format = '¥#,##0.00'
        cell_total.alignment = right_align
        cell_total.font = money_font
        
        ws_summary.cell(row=row, column=6, value="-").alignment = center_align
        
        for col in range(1, 7):
            ws_summary.cell(row=row, column=col).border = border
        
        row += 1
    
    # 总计行
    row_total = row
    ws_summary.cell(row=row_total, column=1, value="").border = border
    ws_summary.cell(row=row_total, column=2, value="合计").font = Font(name='微软雅黑', size=11, bold=True)
    ws_summary.cell(row=row_total, column=2).alignment = center_align
    ws_summary.cell(row=row_total, column=2).border = border
    ws_summary.cell(row=row_total, column=3, value=f"共{total_boxes}个配电箱").alignment = center_align
    ws_summary.cell(row=row_total, column=3).border = border
    ws_summary.cell(row=row_total, column=4, value=sum(len(b['components']) for b in boxes.values())).alignment = center_align
    ws_summary.cell(row=row_total, column=4).border = border
    
    cell_grand_total = ws_summary.cell(row=row_total, column=5, value=total_amount)
    cell_grand_total.number_format = '¥#,##0.00'
    cell_grand_total.font = Font(name='微软雅黑', size=12, bold=True, color='FF0000')
    cell_grand_total.alignment = right_align
    cell_grand_total.border = border
    
    ws_summary.cell(row=row_total, column=6, value="").border = border
    
    # 设置列宽
    ws_summary.column_dimensions['A'].width = 8
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 18
    ws_summary.column_dimensions['D'].width = 12
    ws_summary.column_dimensions['E'].width = 18
    ws_summary.column_dimensions['F'].width = 15
    
    # ========== 详细报价表 ==========
    for box_name, box_info in boxes.items():
        safe_sheet_name = box_name[:31].replace('/', '-').replace('\\', '-')  # Excel表名限制31字符
        ws = wb.create_sheet(title=safe_sheet_name)
        
        # 配电箱标题
        ws['A1'] = f"配电箱：{box_name}"
        ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='1E88E5')
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f"箱体尺寸：{box_info['size']}"
        ws['A2'].font = normal_font
        ws.merge_cells('A2:H2')
        
        # 表头
        detail_headers = ['序号', '元器件名称', '规格型号', '品牌', '数量', '单位', '单价(元)', '小计(元)', '询价来源']
        for col, header in enumerate(detail_headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        # 数据行
        row = 5
        for idx, comp in enumerate(box_info['components'], 1):
            ws.cell(row=row, column=1, value=idx).alignment = center_align
            ws.cell(row=row, column=2, value=comp.get('component_name', '')).alignment = left_align
            ws.cell(row=row, column=3, value=comp.get('spec', '')).alignment = left_align
            ws.cell(row=row, column=4, value=comp.get('brand', '')).alignment = center_align
            ws.cell(row=row, column=5, value=comp.get('quantity', 1)).alignment = center_align
            ws.cell(row=row, column=6, value=comp.get('unit', '个')).alignment = center_align
            
            cell_price = ws.cell(row=row, column=7, value=comp.get('price', 0))
            cell_price.number_format = '¥#,##0.00'
            cell_price.alignment = right_align
            
            cell_subtotal = ws.cell(row=row, column=8, value=comp.get('total', 0))
            cell_subtotal.number_format = '¥#,##0.00'
            cell_subtotal.alignment = right_align
            cell_subtotal.font = money_font
            
            ws.cell(row=row, column=9, value=comp.get('channel', '')).alignment = left_align
            
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = border
            
            row += 1
        
        # 小计行
        box_total = sum(c['total'] for c in box_info['components'])
        ws.cell(row=row, column=1, value="").border = border
        ws.cell(row=row, column=2, value="本配电箱合计").font = Font(name='微软雅黑', size=11, bold=True)
        ws.cell(row=row, column=2).alignment = center_align
        ws.cell(row=row, column=2).border = border
        
        for col in range(3, 8):
            ws.cell(row=row, column=col, value="").border = border
        
        cell_box_total = ws.cell(row=row, column=8, value=box_total)
        cell_box_total.number_format = '¥#,##0.00'
        cell_box_total.font = Font(name='微软雅黑', size=11, bold=True, color='FF0000')
        cell_box_total.alignment = right_align
        cell_box_total.border = border
        
        ws.cell(row=row, column=9, value="").border = border
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 25
    
    # 保存文件
    wb.save(filepath)
    
    return filepath


def generate_simple_excel(data, output_path):
    """
    生成简单版Excel（单个Sheet）
    """
    df = pd.DataFrame(data)
    
    # 重命名列
    column_mapping = {
        'box_name': '配电箱名称',
        'box_size': '箱体尺寸',
        'component_name': '元器件名称',
        'spec': '规格型号',
        'brand': '品牌',
        'quantity': '数量',
        'unit': '单位',
        'price': '单价(元)',
        'total': '小计(元)',
        'channel': '询价来源',
        'url': '询价链接'
    }
    
    df = df.rename(columns=column_mapping)
    
    # 确保列顺序
    ordered_columns = [
        '配电箱名称', '箱体尺寸', '元器件名称', '规格型号', 
        '品牌', '数量', '单位', '单价(元)', '小计(元)', 
        '询价来源', '询价链接'
    ]
    df = df[[col for col in ordered_columns if col in df.columns]]
    
    # 保存
    df.to_excel(output_path, index=False, sheet_name='元器件清单')
    
    return output_path
