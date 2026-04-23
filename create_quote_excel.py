import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "APE-XK报价单"

# 定义样式
header_font = Font(bold=True, size=12)
title_font = Font(bold=True, size=14)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=10, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')

# 标题行
ws['A1'] = "APE-XK配电箱系统完整报价单"
ws['A1'].font = title_font
ws.merge_cells('A1:I1')

# 项目信息
ws['A3'] = "配电箱型号："
ws['B3'] = "APE-XK"
ws['A4'] = "箱体规格："
ws['B4'] = "600×800×200mm，下边距地1.4m，明装"
ws['A5'] = "供电参数："
ws['B5'] = "Pe=10.0kW, Kx=1.0, Pjs=10.0kW, cosΦ=0.80, Ijs=19.0A"

# 表头
headers = ["序号", "元器件名称", "规格型号", "品牌", "数量", "单位", "单价(元)", "小计(元)", "询价来源"]
header_row = 7
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

# 数据
data = [
    # 进线回路
    [1, "隔离开关", "INT 125/63/3P", "施耐德", 2, "只", 120.00, 240.00, "立创商城/海邦电气"],
    [2, "双电源自动转换开关", "WTS-B63/4P", "施耐德万高", 1, "台", 5998.00, 5998.00, "阿里巴巴/海邦电气"],
    [3, "电流互感器", "75/5", "正泰/华邦", 3, "只", 60.00, 180.00, "阿里巴巴"],
    [4, "多功能电力仪表", "KWH+ (RS485通讯)", "安科瑞/中电技术", 1, "台", 265.00, 265.00, "阿里巴巴"],
    # 出线回路
    [5, "微型断路器(1-12路消防设备)", "iC65N-C20A/2P", "施耐德", 12, "只", 58.00, 696.00, "施耐德官网"],
    [6, "微型断路器(第13路备用)", "iC65N-D25A/3P", "施耐德", 1, "只", 85.00, 85.00, "施耐德官网"],
    # 防雷
    [7, "浪涌保护器", "LiGZX4L (40kA)", "利尔德/同为", 1, "套", 380.00, 380.00, "阿里巴巴"],
    # 箱体及辅材
    [8, "配电箱箱体", "600×800×200mm 明装", "基业/中鸿", 1, "台", 240.00, 240.00, "阿里巴巴"],
    [9, "接线端子及辅材", "标准配置", "国产", 1, "套", 150.00, 150.00, "市场参考"],
    [10, "线缆及敷设材料", "NHBV-3×4 CT(JDG25)", "国产", 1, "项", 300.00, 300.00, "市场参考"],
]

# 写入数据
for row_idx, row_data in enumerate(data, header_row + 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx in [5, 6, 7, 8]:  # 数字列居中
            cell.alignment = center_align
        if col_idx in [7, 8]:  # 金额列格式
            cell.number_format = '¥#,##0.00'

# 汇总行
summary_row = header_row + len(data) + 1
ws.cell(row=summary_row, column=1, value="").border = thin_border
ws.cell(row=summary_row, column=2, value="设备材料合计").border = thin_border
ws.cell(row=summary_row, column=3, value="").border = thin_border
ws.cell(row=summary_row, column=4, value="").border = thin_border
ws.cell(row=summary_row, column=5, value="").border = thin_border
ws.cell(row=summary_row, column=6, value="").border = thin_border
ws.cell(row=summary_row, column=7, value="").border = thin_border
cell = ws.cell(row=summary_row, column=8, value=8534.00)
cell.border = thin_border
cell.number_format = '¥#,##0.00'
cell.font = Font(bold=True)
ws.cell(row=summary_row, column=9, value="").border = thin_border

ws.cell(row=summary_row+1, column=2, value="人工安装费(约15%)").border = thin_border
for col in range(3, 9):
    ws.cell(row=summary_row+1, column=col, value="").border = thin_border
cell = ws.cell(row=summary_row+1, column=8, value=1280.10)
cell.border = thin_border
cell.number_format = '¥#,##0.00'
ws.cell(row=summary_row+1, column=9, value="").border = thin_border

ws.cell(row=summary_row+2, column=2, value="报价总计").border = thin_border
ws.cell(row=summary_row+2, column=2).font = Font(bold=True, size=12)
for col in range(3, 9):
    ws.cell(row=summary_row+2, column=col, value="").border = thin_border
cell = ws.cell(row=summary_row+2, column=8, value=9814.10)
cell.border = thin_border
cell.number_format = '¥#,##0.00'
cell.font = Font(bold=True, size=12)
cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
ws.cell(row=summary_row+2, column=9, value="").border = thin_border

# 关键说明
note_row = summary_row + 5
ws.cell(row=note_row, column=1, value="关键元器件说明：").font = Font(bold=True, size=11)
ws.cell(row=note_row+1, column=1, value="1. 双电源自动转换开关WTS-B63/4P：施耐德万高品牌，4P,63A，智能控制器，价格¥5,998-12,680")
ws.cell(row=note_row+2, column=1, value="2. 多功能电力仪表KWH+：RS485通讯，测量电压/电流/功率/电能等参数，价格¥265-350")
ws.cell(row=note_row+3, column=1, value="3. 浪涌保护器LiGZX4L：40kA，4P，用于防雷保护，价格¥350-380")
ws.cell(row=note_row+4, column=1, value="注：本报价单价格有效期为7天，实际价格以供应商报价为准。")

# 设置列宽
column_widths = [8, 28, 22, 18, 8, 8, 14, 14, 25]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# 保存
wb.save("配电箱自动组价系统/APE-XK报价单_完整版.xlsx")
print("Excel报价单已生成: 配电箱自动组价系统/APE-XK报价单_完整版.xlsx")
