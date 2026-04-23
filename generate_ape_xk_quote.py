#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
APE-XK配电箱报价单生成脚本
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "APE-XK报价单"

# 定义样式
header_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
title_font = Font(name='微软雅黑', size=12, bold=True)
normal_font = Font(name='微软雅黑', size=10)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
alt_fill = PatternFill(start_color='D6DCE5', end_color='D6DCE5', fill_type='solid')
total_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 标题行
ws['A1'] = 'APE-XK配电箱报价单'
ws['A1'].font = Font(name='微软雅黑', size=16, bold=True)
ws.merge_cells('A1:J1')
ws['A1'].alignment = center_align

# 基本信息
info_data = [
    ['配电箱名称', 'APE-XK', '项目名称', '厂房项目配电箱'],
    ['箱体尺寸', '600×800×200mm', '安装方式', '明装，下边距地1.4m'],
    ['报价日期', '2026年4月17日', '', '']
]

row = 3
for info in info_data:
    ws[f'A{row}'] = info[0]
    ws[f'B{row}'] = info[1]
    ws[f'C{row}'] = info[2]
    ws[f'D{row}'] = info[3]
    ws[f'A{row}'].font = title_font
    ws[f'C{row}'].font = title_font
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].border = thin_border
        ws[f'{col}{row}'].alignment = center_align
    row += 1

row += 1

# 配电箱参数
ws[f'A{row}'] = '配电箱参数'
ws[f'A{row}'].font = title_font
ws.merge_cells(f'A{row}:D{row}')
row += 1

params = [
    ['参数', 'Pe（额定功率）', 'Kx（需要系数）', 'Pjs（计算功率）', 'cosΦ（功率因数）', 'Ijs（计算电流）'],
    ['数值', '10.0kW', '1.0', '10.0kW', '0.80', '19.0A']
]

for i, param_row in enumerate(params):
    for col_idx, value in enumerate(param_row, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align
        if i == 0:
            cell.font = title_font
            cell.fill = header_fill
            cell.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    row += 1

row += 1

# 元器件清单表头
ws[f'A{row}'] = '元器件清单及报价'
ws[f'A{row}'].font = title_font
ws.merge_cells(f'A{row}:J{row}')
row += 1

headers = ['序号', '元器件名称', '规格型号', '品牌', '数量', '单位', '单价(元)', '小计(元)', '询价来源', '询价链接']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col_idx, value=header)
    cell.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align
row += 1

# 元器件数据
components = [
    [1, '进线隔离开关', 'INT 125/63/3P', '施耐德', 2, '个', 563.10, 1126.20, '千思买工品仓', 'https://www.gpcang.com/Goods/20441.html'],
    [2, '双电源开关', 'WTS-863/4P', '施耐德万高', 1, '个', 3500.00, 3500.00, '海邦电气', 'https://www.haibangdq.com/product/detail/150244271'],
    [3, '电流互感器', '75/5A', '施耐德', 1, '个', 180.00, 180.00, '工控猫商城', 'https://www.gkmao.com'],
    [4, '多功能电力仪表', 'DM2350N', '施耐德', 1, '个', 650.00, 650.00, '工控猫商城', 'https://www.gkmao.com'],
    [5, '消防出线断路器', 'iC65N-C20A/2P', '施耐德', 12, '个', 75.00, 900.00, '工控猫商城', 'https://www.gkmao.com'],
    [6, '备用断路器', 'iC65N-D25A/3P', '施耐德', 1, '个', 120.00, 120.00, '工控猫商城', 'https://www.gkmao.com'],
    [7, '防雷模块(SPD)', 'LiD40X4 40kA', '施耐德', 1, '个', 350.00, 350.00, '工控猫商城', 'https://www.gkmao.com'],
]

start_row = row
for i, comp in enumerate(components):
    fill = alt_fill if i % 2 == 1 else None
    for col_idx, value in enumerate(comp, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align if col_idx != 10 else left_align
        if fill:
            cell.fill = fill
        if col_idx in [7, 8]:  # 金额列右对齐
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.number_format = '¥#,##0.00'
    row += 1

# 总计行
total_row = row
ws[f'A{total_row}'] = ''
ws.merge_cells(f'A{total_row}:F{total_row}')
ws[f'G{total_row}'] = '合计'
ws[f'G{total_row}'].font = title_font
ws[f'G{total_row}'].alignment = center_align
ws[f'G{total_row}'].border = thin_border
ws[f'G{total_row}'].fill = total_fill

ws[f'H{total_row}'] = 6726.20
ws[f'H{total_row}'].font = Font(name='微软雅黑', size=12, bold=True)
ws[f'H{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
ws[f'H{total_row}'].number_format = '¥#,##0.00'
ws[f'H{total_row}'].border = thin_border
ws[f'H{total_row}'].fill = total_fill

ws.merge_cells(f'I{total_row}:J{total_row}')
ws[f'I{total_row}'].border = thin_border
ws[f'I{total_row}'].fill = total_fill

row += 2

# 询价渠道
ws[f'A{row}'] = '询价渠道说明'
ws[f'A{row}'].font = title_font
ws.merge_cells(f'A{row}:D{row}')
row += 1

channels = [
    ['优先级', '渠道名称', '网址'],
    ['1', '电小二官网（电气元器件询价）', 'https://www.dianxiaoer.com'],
    ['2', 'ABB直通车', 'https://mall.abb.com.cn'],
    ['3', '天工直通车', 'https://www.titanmatrix.com'],
    ['4', '工控猫商城', 'https://www.gkmao.com'],
]

for i, channel in enumerate(channels):
    for col_idx, value in enumerate(channel, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align
        if i == 0:
            cell.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
            cell.fill = header_fill
    row += 1

row += 1

# 备注
ws[f'A{row}'] = '备注'
ws[f'A{row}'].font = title_font
ws.merge_cells(f'A{row}:J{row}')
row += 1

remarks = [
    '1. 本报价为参考价，实际采购价格以供应商实时报价为准',
    '2. 消防电源监测模块采用安科瑞品牌（国产一线），施耐德/ABB无同规格产品',
    '3. 多功能电力仪表采用施耐德DM2350N系列（常规参数配置）',
    '4. 电流互感器采用施耐德MET SEC系列75/5A变比',
    '5. 防雷模块采用施耐德Easy9系列40kA家用浪涌保护器',
    '6. 箱体规格：600×800×200mm（箱体价格未计入）',
]

for remark in remarks:
    ws[f'A{row}'] = remark
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].alignment = left_align
    row += 1

row += 2
ws[f'A{row}'] = '报价人：李冠亚的扣子助手'
ws[f'A{row}'].font = normal_font
ws[f'D{row}'] = '报价时间：2026年4月17日'
ws[f'D{row}'].font = normal_font

# 设置列宽
column_widths = {
    'A': 8, 'B': 18, 'C': 18, 'D': 10, 'E': 8, 'F': 8, 'G': 12, 'H': 12, 'I': 15, 'J': 45
}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 保存文件
output_path = 'APE-XK报价单_20260417_v2.xlsx'
wb.save(output_path)
print(f"报价单已生成: {output_path}")
